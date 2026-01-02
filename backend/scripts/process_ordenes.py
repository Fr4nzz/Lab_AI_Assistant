#!/usr/bin/env python3
"""
Process ordenes XLSX to generate curated orders list for fuzzy search.

Usage:
    python process_ordenes.py <input_xlsx> [output_csv]

Example:
    python process_ordenes.py ../ordenes-2026-01-01-17-15-41.xlsx
    python process_ordenes.py ../ordenes-2026-01-01-17-15-41.xlsx lista_de_ordenes.csv

This script:
1. Reads the ordenes XLSX (header on row 4)
2. Extracts key fields: order_num, date, patient_name, cedula, exams, total
3. Outputs a clean CSV for fuzzy search
"""
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)


def process_ordenes(input_path: str, output_path: str = None):
    """
    Process ordenes XLSX and generate curated orders list.

    Args:
        input_path: Path to input ordenes XLSX
        output_path: Path to output CSV (default: backend/config/lista_de_ordenes.csv)
    """
    input_file = Path(input_path)
    if not input_file.exists():
        print(f"Error: Input file not found: {input_path}")
        sys.exit(1)

    # Default output path
    if output_path is None:
        output_path = Path(__file__).parent.parent / "config" / "lista_de_ordenes.csv"
    else:
        output_path = Path(output_path)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {input_file}")

    # Load workbook
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active

    # Get header row (row 4, 0-indexed = row 3)
    header_row = 4
    headers = []
    for cell in ws[header_row]:
        headers.append(cell.value if cell.value else "")

    # Find column indices
    col_map = {
        'order_num': headers.index('Orden No.') if 'Orden No.' in headers else -1,
        'fecha': headers.index('Fecha orden') if 'Fecha orden' in headers else -1,
        'nombres': headers.index('Nombres') if 'Nombres' in headers else -1,
        'apellidos': headers.index('Apellidos') if 'Apellidos' in headers else -1,
        'cedula': headers.index('Identificación') if 'Identificación' in headers else -1,
        'examenes': headers.index('Exámenes') if 'Exámenes' in headers else -1,
        'total': headers.index('Total') if 'Total' in headers else -1,
    }

    # Validate required columns
    missing = [k for k, v in col_map.items() if v == -1]
    if missing:
        print(f"Warning: Missing columns: {missing}")

    orders = []
    row_count = 0

    # Process data rows (starting from row 5)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_count += 1

        # Get values safely
        order_num = str(row[col_map['order_num']] or '').strip() if col_map['order_num'] >= 0 else ''
        if not order_num:
            continue

        fecha = str(row[col_map['fecha']] or '').strip() if col_map['fecha'] >= 0 else ''
        nombres = str(row[col_map['nombres']] or '').strip() if col_map['nombres'] >= 0 else ''
        apellidos = str(row[col_map['apellidos']] or '').strip() if col_map['apellidos'] >= 0 else ''
        cedula = str(row[col_map['cedula']] or '').strip() if col_map['cedula'] >= 0 else ''
        examenes = str(row[col_map['examenes']] or '').strip() if col_map['examenes'] >= 0 else ''
        total = str(row[col_map['total']] or '').strip() if col_map['total'] >= 0 else ''

        # Combine names
        patient_name = f"{apellidos}, {nombres}".strip(', ')

        orders.append({
            'order_num': order_num,
            'fecha': fecha,
            'patient_name': patient_name,
            'nombres': nombres,
            'apellidos': apellidos,
            'cedula': cedula,
            'examenes': examenes,
            'total': total,
        })

    wb.close()

    print(f"Processed {row_count} rows, found {len(orders)} orders")

    # Sort by order number descending (most recent first)
    orders.sort(key=lambda x: x['order_num'], reverse=True)

    # Write output CSV
    fieldnames = ['order_num', 'fecha', 'patient_name', 'nombres', 'apellidos', 'cedula', 'examenes', 'total']

    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for order in orders:
            writer.writerow(order)

    print(f"Output: {output_path}")
    print(f"Total orders: {len(orders)}")

    # Print summary by month
    months = {}
    for order in orders:
        fecha = order['fecha']
        if fecha:
            try:
                # Parse date (format: DD/MM/YYYY)
                if '/' in fecha:
                    parts = fecha.split('/')
                    if len(parts) == 3:
                        month_key = f"{parts[2]}-{parts[1]}"  # YYYY-MM
                        months[month_key] = months.get(month_key, 0) + 1
            except Exception:
                pass

    if months:
        print("\nOrders by month:")
        for month, count in sorted(months.items(), reverse=True)[:12]:
            print(f"  {month}: {count}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python process_ordenes.py <input_xlsx> [output_csv]")
        print("Example: python process_ordenes.py ../../ordenes-2026-01-01-17-15-41.xlsx")
        sys.exit(1)

    input_xlsx = sys.argv[1]
    output_csv = sys.argv[2] if len(sys.argv) > 2 else None

    process_ordenes(input_xlsx, output_csv)
